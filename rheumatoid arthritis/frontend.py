from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from datetime import date
from tkinter import filedialog
from PIL import Image ,ImageTk
import os 
from tkinter.ttk import Combobox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib
import pandas as pd
bgg="#06283D"
fgbg="#EDEDED"
ffg="#06283D"
file =pathlib.Path('PredictionData.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="ID"
    sheet['B1']="Gender"     #Male or female
    sheet['C1']="ACPA_AntiBodies"      #Yes or no
    sheet['D1']="arthralgia" #Yes or no
    sheet['E1']="rheumatoidfactorAntiBodies"
    sheet['F1']="siblings"
    sheet['G1']="ParentOfRA"
    sheet['H1']="smoker"
    sheet['I1']="HLA-DRB gene"
    sheet['J1']="amino_acids_at_positions_11,13,71,74"
    sheet['K1']="AntiCCP"
    sheet['L1']="Result"
    sheet['M1']="NumberOfAccess"


    file.save('PredictionData.xlsx')
    file.close()

file =pathlib.Path('DiagnmoseData.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="ID"
    sheet['B1']="Number Of Large Joint"    
    sheet['C1']="Number Of Small Joint"      #Yes or no
    sheet['D1']="Totall number of joints" #Yes or no
    sheet['E1']="Result of Rf"
    sheet['F1']="Result of ACPA"
    sheet['G1']="symptoms duration"
    sheet['H1']="C-reactive protien"
    sheet['I1']="result of ESR"
    sheet['J1']="Final result"
    sheet['K1']="NumberOfAccess"


    file.save('DiagnmoseData.xlsx')
    file.close()

root =Tk()
root.title("Romatoide")
root.geometry("925x500+300+200")
root.configure(bg="#fff")
root.resizable(False,False)

radio=IntVar()     # Gender
radio1=StringVar() # ang
radio2=StringVar() # ACPA
radio3=StringVar() # PRA
radio4=StringVar() # SRA
radio5=StringVar() # Smoker
radio6=StringVar() # GEN
radio7=StringVar() # AminoACid

radio8=StringVar()  #ACCP

radio9=StringVar()  #RF


#----------------------------------------Diagnose

radio10=StringVar() # Number Of Large Joint
radio11=StringVar() # Number Of Small Joint
radio12=StringVar() # Totall number of joints
radio13=StringVar() # Result of Rf
radio14=StringVar() # Result of ACPA
radio15=StringVar() # Symp Duration
radio16=StringVar() # C Reactive prot
radio17=StringVar() # ESR




def Save():
    try:
        B1=Gender
        C1=ACPA
        D1=ang
        E1=PRA
        F1=SRA
        G1=Smoker
        H1=GEN
        I1=ACCP
        J1=RF
    except:
        messagebox.showerror("Missing data")


    file=openpyxl.load_workbook('PredictionData.xlsx')
    sheet=file.active    
    sheet.cell(column=2 + 12* Num,row=UserRow+2,value=Gender)
    sheet.cell(column=4 +12* Num,row=UserRow+2,value= ang)
    sheet.cell(column=3+12*Num,row=UserRow+2,value= ACPA)
    sheet.cell(column=5+12*Num,row=UserRow+2,value=RF)
    sheet.cell(column=7+12*Num,row=UserRow+2,value= PRA)
    sheet.cell(column=6+12*Num,row=UserRow+2,value= SRA)
    sheet.cell(column=8+12*Num,row=UserRow+2,value= Smoker)
    sheet.cell(column=9+12*Num,row=UserRow+2,value= GEN)
    sheet.cell(column=11+12*Num,row=UserRow+2,value=ACCP)

    sheet.cell(column=13,row=UserRow+2,value=Num+1)


    file.save('PredictionData.xlsx')
    PScreen.destroy()

def Selection():
    global Gender
    
    value=radio.get()
    
    if(value==1):
        Gender="Male"
    else:
        Gender="Female"

def angfn():
    global ang
    ang=radio1.get()

def ACPAfn():
    global ACPA
    ACPA=radio2.get()


def PRAfn():
    global PRA
    PRA=radio3.get()

def Smokerfn():
    global Smoker
    Smoker=radio5.get()

def Genfn():
    global GEN
    GEN=radio6.get()

def SRAfn():
    global SRA
    SRA=radio4.get()


def ACCPFn():
    global ACCP
    ACCP=radio8.get()


def RFFn():
    global RF
    RF=radio9.get()

#------------------------------------------------------diagnose fn 

def SaveD():
    try:
        B1=LGP
        C1=SGP
        D1=TNOJ
        E1=DRF
        F1=ACPA
        G1=SD
        H1=CRP
        I1=ESR
    except:
        messagebox.showerror("Missing data")

    print("mody")
    file=openpyxl.load_workbook('DiagnmoseData.xlsx')
    sheet=file.active    
    sheet.cell(column=2 + 10* NumD,row=UserRow+2,value=LGP)
    sheet.cell(column=3 +10* NumD,row=UserRow+2,value= SGP)
    sheet.cell(column=4+10*NumD,row=UserRow+2,value= TNOJ)
    sheet.cell(column=5+10*NumD,row=UserRow+2,value=DRF)
    sheet.cell(column=6+10*NumD,row=UserRow+2,value= ACPA)
    sheet.cell(column=7+10*NumD,row=UserRow+2,value= SD)
    sheet.cell(column=8+10*NumD,row=UserRow+2,value= CRP)
    sheet.cell(column=9+10*NumD,row=UserRow+2,value= ESR)

    sheet.cell(column=11,row=UserRow+2,value=NumD+1)


    file.save('DiagnmoseData.xlsx')
    file.close()
    DScreen.destroy()


def LGPFN():
    global LGP
    LGP=radio10.get()


def SGPFN():
    global SGP
    SGP=radio11.get()

def TNOJFN():
    global TNOJ
    TNOJ=radio12.get()

def DRFFN():
    global DRF
    DRF=radio13.get()


def ACPAFN():
    global ACPA
    ACPA=radio14.get()


def SDFN():
    global SD
    SD=radio15.get()

def CRPFN():
    global CRP
    CRP=radio16.get()

def ESRFN():
    global ESR
    ESR =radio17.get()





    



def Predication_ButtonFn():
    global PScreen
    PScreen=Toplevel(root)
    PScreen.title( "Prediction" )
    #Screen.geometry('925x500 +300 +200')
    PScreen.config(bg='#06283D')
    Label(PScreen,text='Welcome ',fg='#57a178',bg='#06283D',font=('Calibri(Body)',50,'bold')).pack()

    Button(PScreen,width=30,pady=7,text='Predict',bg='white',fg='#06283D',border=0,font=('Calibri(Body)',30,'bold'),command=Save).place(x=45,y=700)

    
    obj=LabelFrame(PScreen,text="Select Which You Have",font=20,bd=2,width=3000,bg='Purple',fg='#EDEDED',height=500,relief=GROOVE)
    obj.place(x=0,y=100)

    #Gender
    Label(obj,width=5,height=1,text='Gender:',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=20)
    R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg="black",fg='LightGreen',command=Selection )
    R1.place(x=50,y=20)

    R2=Radiobutton(obj,text="FeMale",variable=radio,value=2,bg="black",fg='LightGreen',command=Selection )
    R2.place(x=110,y=20)
    


    #ACPA
    Label(obj,width=15,height=1,text='ACPA:',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=60)
    R5=Radiobutton(obj,text="Greater than 20",variable=radio2,value="Greater than 20",bg="black",fg='LightGreen',command=ACPAfn )
    R5.place(x=115,y=59)
    R6=Radiobutton(obj,text="less than or equal 20",variable=radio2,value="less than or equal 20",bg="black",fg='LightGreen',command=ACPAfn )
    R6.place(x=230,y=59)


    #arthralgia
    Label(obj,width=15,height=1,text='arthralgia:',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=90)
    R3=Radiobutton(obj,text="YES",variable=radio1,value="YES",bg="black",fg='LightGreen',command=angfn )
    R3.place(x=115,y=90)
    R4=Radiobutton(obj,text="NO",variable=radio1,value="NO",bg="black",fg='LightGreen',command=angfn )
    R4.place(x=190,y=90)
    
    #Parent of RA
    Label(obj,width=35,height=1,text='Does a parent have Rheumatoide-Arthritis?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=120)
    R7=Radiobutton(obj,text="YES",variable=radio3,value="YES",bg="black",fg='LightGreen',command=PRAfn )
    R7.place(x=255,y=120)
    R8=Radiobutton(obj,text="NO",variable=radio3,value="NO",bg="black",fg='LightGreen',command=PRAfn )
    R8.place(x=315 ,y=120)


    #smoker
    Label(obj,width=15,height=1,text='Are you smoking',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=150)
    R7=Radiobutton(obj,text="YES",variable=radio5,value="YES",bg="black",fg='LightGreen',command=Smokerfn )
    R7.place(x=115,y=150)
    R8=Radiobutton(obj,text="NO",variable=radio5,value="NO",bg="black",fg='LightGreen',command=Smokerfn )
    R8.place(x=190 ,y=150)



    #Gen
    Label(obj,width=35,height=1,text='Do you have the HLA-DRB1(HLA-SR)Gene?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=180)
    R7=Radiobutton(obj,text="YES",variable=radio6,value="YES",bg="black",fg='LightGreen',command=Genfn )
    R7.place(x=255,y=180)
    R8=Radiobutton(obj,text="NO",variable=radio6,value="NO",bg="black",fg='LightGreen',command=Genfn )
    R8.place(x=315 ,y=180)

    # SRA   
    Label(obj,width=35,height=1,text='Is one of your sibling have Rheumatoide-Arthritis ?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=210)
    R18=Radiobutton(obj,text="YES",variable=radio4,value="YES",bg="black",fg='LightGreen',command=SRAfn )
    R18.place(x=255,y=210)
    R19=Radiobutton(obj,text="NO",variable=radio4,value="NO",bg="black",fg='LightGreen',command=SRAfn )
    R19.place(x=315 ,y=210)
    R20=Radiobutton(obj,text="TWIN",variable=radio4,value="TWIN",bg="black",fg='LightGreen',command=SRAfn )
    R20.place(x=375 ,y=210)


    # AMino ACid
    Label(obj,width=35,height=1,text='Do you have amino acid at position?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=240)
    R21=Radiobutton(obj,text="YES",variable=radio7,value="YES",bg="black",fg='LightGreen',command=Genfn )
    R21.place(x=255,y=240)
    R22=Radiobutton(obj,text="NO",variable=radio7,value="NO",bg="black",fg='LightGreen',command=Genfn )
    R22.place(x=315 ,y=240)


    # anticcp 

    Label(obj,width=35,height=1,text='Do you have Anti-CCP Anti bodies?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=270)
    R23=Radiobutton(obj,text="ACCP1",variable=radio8,value="YES",bg="black",fg='LightGreen',command=ACCPFn )
    R23.place(x=255,y=270)
    R24=Radiobutton(obj,text="ACCP2",variable=radio8,value="ACCP2",bg="black",fg='LightGreen',command=ACCPFn )
    R24.place(x=340 ,y=270)

    R25=Radiobutton(obj,text="BOTH",variable=radio8,value="BOTH",bg="black",fg='LightGreen',command=ACCPFn )
    R25.place(x=430,y=270)
    R26=Radiobutton(obj,text="NONE",variable=radio8,value="NONE",bg="black",fg='LightGreen',command=ACCPFn )
    R26.place(x=520 ,y=270)

    #RF
    Label(obj,width=15,height=1,text='What is the value of Rheumatoid factor',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=300)
    R5=Radiobutton(obj,text="Greater than 20",variable=radio9,value="Greater than 20",bg="black",fg='LightGreen',command=RFFn )
    R5.place(x=115,y=300)
    R6=Radiobutton(obj,text="less than or equal 20",variable=radio9,value="less than or equal 20",bg="black",fg='LightGreen',command=RFFn )
    R6.place(x=230,y=300)




    

def Diagnose_ButtonFn():
    global DScreen
    DScreen=Toplevel(root)
    DScreen.title( "Diagnose" )
    #Screen.geometry('925x500 +300 +200')
    DScreen.config(bg='#06283D')
    
    Label(DScreen,text='Welcome ',fg='#57a178',bg='white',font=('Calibri(Body)',50,'bold')).pack()

    Button(DScreen,width=30,pady=7,text='Diagnose',bg='white',fg='#06283D',border=0,font=('Calibri(Body)',30,'bold'),command=SaveD).place(x=45,y=700)

    
    obj=LabelFrame(DScreen,text="Select Which You Have",font=20,bd=2,width=3000,bg='Purple',fg='#EDEDED',height=500,relief=GROOVE)
    obj.place(x=0,y=100)


    # large joints pain
    Label(obj,width=35,height=1,text='How many large joints pain?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=20)
    D1=Radiobutton(obj,text="1",variable=radio10,value="1",bg="black",fg='LightGreen',command=LGPFN )
    D1.place(x=255,y=20)
    D2=Radiobutton(obj,text="Greater Than 1",variable=radio10,value="Greater Than 1",bg="black",fg='LightGreen',command=LGPFN )
    D2.place(x=400 ,y=20)


    # small joints pain
    Label(obj,width=35,height=1,text='How many large joints pain?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=50)
    D3=Radiobutton(obj,text="1-3",variable=radio11,value="1-3",bg="black",fg='LightGreen',command=SGPFN )
    D3.place(x=255,y=50)
    D4=Radiobutton(obj,text="4-10",variable=radio11,value="4-10",bg="black",fg='LightGreen',command=SGPFN )
    D4.place(x=400 ,y=50)

     # Total Number of joints
    Label(obj,width=40,height=1,text='What is the total number of joints pain ?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=80)
    D5=Radiobutton(obj,text="Greater than 10",variable=radio12,value="Greater than 10",bg="black",fg='LightGreen',command=TNOJFN )
    D5.place(x=300,y=80)
    D6=Radiobutton(obj,text="less than or equal 10",variable=radio12,value="less than or equal 10",bg="black",fg='LightGreen',command=TNOJFN )
    D6.place(x=500 ,y=80)


    # rheumatoid factor
    Label(obj,width=40,height=1,text='What is the result of rheumatoid factor?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=110)
    D7=Radiobutton(obj,text="Less than or equal 20",variable=radio13,value="Less than or equal 20",bg="black",fg='LightGreen',command=DRFFN )
    D7.place(x=300,y=110)
    D8=Radiobutton(obj,text="From 20 to 23 than or equal 20",variable=radio13,value="From 20 to 23 than or equal 20",bg="black",fg='LightGreen',command=DRFFN )
    D8.place(x=500 ,y=110)
    D9=Radiobutton(obj,text="Greater than 23",variable=radio13,value="Greater than 23",bg="black",fg='LightGreen',command=DRFFN )
    D9.place(x=700 ,y=110)


    # ACPA factor
    Label(obj,width=35,height=1,text='What is the result of ACPA factor?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=140)
    D10=Radiobutton(obj,text="Less than or equal 20",variable=radio14,value="Less than or equal 20",bg="black",fg='LightGreen',command=ACPAFN )
    D10.place(x=255,y=140)
    D11=Radiobutton(obj,text="From 20 to 23 than or equal 20",variable=radio14,value="From 20 to 23 than or equal 20",bg="black",fg='LightGreen',command=ACPAFN )
    D11.place(x=400 ,y=140)
    D12=Radiobutton(obj,text="Greater than 23",variable=radio14,value="Greater than 23",bg="black",fg='LightGreen',command=ACPAFN )
    D12.place(x=600 ,y=140)


    # Symptimous 
    Label(obj,width=35,height=1,text='What is the Symptimous Duration?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=170)
    D13=Radiobutton(obj,text="Less than 6 weeks",variable=radio15,value="Less than 6 weeks",bg="black",fg='LightGreen',command=SDFN )
    D13.place(x=255,y=170)
    D14=Radiobutton(obj,text="Greater than or equal 6 weeks",variable=radio15,value="Greater than or equal 6 weeks",bg="black",fg='LightGreen',command=SDFN )
    D14.place(x=400 ,y=170)



     # CRP 
    Label(obj,width=35,height=1,text='Is your Result of CRP (c- reactive protein) less than 6?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=200)
    D13=Radiobutton(obj,text="Yes",variable=radio16,value="Yes",bg="black",fg='LightGreen',command=CRPFN )
    D13.place(x=255,y=200)
    D14=Radiobutton(obj,text="No",variable=radio16,value="No",bg="black",fg='LightGreen',command=CRPFN )
    D14.place(x=400 ,y=200)



     # ESR 
    Label(obj,width=35,height=1,text='Is your result of ESR in range 1 - 15?',fg='#57a178',bg='#06283D',font=('Calibri(Body)',9,'bold')).place(x=0,y=230)
    D13=Radiobutton(obj,text="Yes",variable=radio17,value="Yes",bg="black",fg='LightGreen',command=ESRFN )
    D13.place(x=255,y=230)
    D14=Radiobutton(obj,text="No",variable=radio17,value="No",bg="black",fg='LightGreen',command=ESRFN )
    D14.place(x=400 ,y=230)
   







   

    
    DScreen.mainloop()


def SignIn_ButtonFn():
    passs=password.get()
    data= pd.read_excel("PredictionData.xlsx")
    for i in range(len(data)):
        if int(data.iloc[i,0]) == int(passs):
            global Num     #Number that the user access the Prediction
            global NumD    #Number that the user access the diagnose
            global UserRow
            Num=data.iloc[i,12]
            
            UserRow=i
            dataD= pd.read_excel("DiagnmoseData.xlsx")
            NumD= dataD.iloc[i,10]

            
            Screen=Toplevel(root)
            Screen.title( "Diagnose")
            Screen.config(bg='white')
            Label(Screen,text='Choose either Diagnosed Or predicated ',fg='#57a178',bg='white',font=('Calibri(Body)',20,'bold')).pack()
            Label(Screen,text='please, read the awarness section before you choose one',fg='#57a178',bg='white',font=('Calibri(Body)',20,'bold')).place(x=500,y=200)

            PredicationButton=Button(Screen,width=39,pady=7,text='PredicationButton',bg='green',fg='purple',font=('Calibri(Heading)',10),border=10,command=Predication_ButtonFn).place(x=595,y=300)

            DiagnoseButton=Button(Screen,width=39,pady=7,text='DiagnoseButton',bg='green',fg='purple',font=('Calibri(Heading)',10),border=10,command=Diagnose_ButtonFn ).place(x=600,y=400)
            img =PhotoImage(file="Diganose1.png")
            Label(Screen,image=img,bg='white').place(x=950,y=392)

            Screen.mainloop()
            return
    
    messagebox.showerror("Invalid","Wrong Id")



def SignUp():
    Screen=Toplevel(root)
    Screen.title( "SignUp")
    Screen.config(bg='purple')

    Label(Screen,width=25,text="Please choose Your Id",fg='black',bg='white',font=('Microsoft YaHei UI Light',8)).place(x=0,y=120)
    global Id
    Id=Entry(Screen,width=25,fg='white',border=0,bg='Green',font=('Microsoft YaHei UI Light',11))
    Id.place(x=180,y=122)
    print(Id)
    Button(Screen,width=30,pady=7,text='Sign Up',bg='black',fg='white',border=0,command=SignUp1).place(x=300,y=235)
    Screen.mainloop()

def SignUp1():
    id= int(Id.get() )
    data= pd.read_excel("PredictionData.xlsx")
    for i in range(len(data)):
        if int(data.iloc[i,0]) == id:
            messagebox.showerror("Invalid","This Id Is taken Please choose another one")
            return
    
    file=openpyxl.load_workbook('PredictionData.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=str (id))
    sheet.cell(column=13,row=sheet.max_row,value=str (0) )
    file.save("PredictionData.xlsx")
    file.close()

    file=openpyxl.load_workbook('DiagnmoseData.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=str (id))
    sheet.cell(column=11,row=sheet.max_row,value=str (0) )

    file.save("DiagnmoseData.xlsx")
    file.close()


    messagebox.showinfo("valid","Done!")
    

    

img =PhotoImage(file="Rom.png")
Label(root,image=img,bg='white').place(x=50,y=50)
heading=Label(root,text='Welcome ',fg='Black',bg='white',font=('Microsoft YaHei UI Light',23,'bold') )
heading.place(x=400,y=5)

frame=Frame(root,width=350,heigh=350,bg='white')
frame.place(x=480,y=70)

heading1=Label(frame,text='Sign in',fg='Black',bg='white',font=('Microsoft YaHei UI Light',20,'bold'))
heading1.place(x=125,y=10)
#-----------------------------------------------------------




#-----------------------------------------------------------
def On_EnterP(e):
    password.delete(0,'end')

def On_LeaveP(e):
    pp=password.get()
    if pp=='':
        password.insert(0,'Id')
password=Entry(frame,width=25,fg='black',border=0,bg='white',font=('Microsoft YaHei UI Light',11,))
password.place(x=50,y=120)
password.insert(0,'Id')

password.bind('<FocusIn>',On_EnterP )
password.bind('<FocusOut>',On_LeaveP )
Frame(frame,width=228,height=2,bg='black').place(x=50,y=140)


Button(frame,width=30,pady=7,text='Sign in',bg='#57a1f8',fg='white',border=0,command=SignIn_ButtonFn).place(x=45,y=235)

label=Label(frame,text="Don't have an account?",fg='black',bg='white',font=('Microsoft YaHei UI Light',9)).place(x=45,y=280)

Sign_up=Button(frame,width=15,text="Sign up",border=0,bg='white',fg='#57a1f8',command=SignUp).place(x=180,y=281)



root.mainloop()